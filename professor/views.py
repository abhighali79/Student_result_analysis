from django.http import FileResponse
import openpyxl
from users.models import Branch, Batch, Student
from users.models import CustomUser
from student.models import marks,Document
from django.shortcuts import render, redirect
from django.contrib.auth.models import User
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from users.models import CustomUser,Batch,Branch
from users.forms import BranchForm, BatchForm
from PIL import Image
import matplotlib
matplotlib.use('Agg')  # Non-GUI backend
import matplotlib.pyplot as plt
from openpyxl.styles import Alignment
import numpy as np

import openpyxl.styles.alignment 
from openpyxl.styles import Border, Side, Alignment
 
@login_required
def professor_dashboard(request):
    """Render the professor dashboard with branches, batches, and sems."""
    emp_id=request.session['empid'] 
    if not request.user.is_professor:
        messages.error(request, 'Please log in as a professor.')
        return redirect('professor_login')
    if not request.user.is_authenticated:
        return redirect('professor_login')
    if not emp_id:
        messages.error(request, 'Session expired. Please log in again.')
        return redirect('professor_login')
    if request.method == 'POST' and 'branch' in request.POST:
        if Branch.objects.filter(branch=request.POST['branch']).exists():
            messages.error(request, 'Branch already exists.')
            return redirect('professor_dashboard')
        form = BranchForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Branch added successfully.')
            return redirect('professor_dashboard')
        
    if request.method == 'POST' and 'batch' in request.POST:
        if Batch.objects.filter(batch=request.POST['batch']).exists():
            messages.error(request, 'Batch already exists.')
            return redirect('professor_dashboard')
        form = BatchForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, 'Batch added successfully.')
            return redirect('professor_dashboard')
    


    branches = Branch.objects.all()
    branch_stud={}
    for branch in branches:
        count=Student.objects.filter(branch_id=branch.id).count()
        branch_stud[branch.branch]=count
    print(branch_stud)
    form1= BranchForm()
    form2= BatchForm()
    context = {'branches': branches,'branch_stud':branch_stud,'form1':form1,'form2':form2} 
    
    return render(request, 'a.html', context)

@login_required
def batch(request, branch_id):
    emp_id=request.session['empid'] 
    if not request.user.is_professor:
        messages.error(request, 'Please log in as a professor.')
        return redirect('professor_login')
    if not request.user.is_authenticated:
        return redirect('professor_login')
    if not emp_id:
        messages.error(request, 'Session expired. Please log in again.')
        return redirect('professor_login')
    print(f"Branch ID: {branch_id}")
    # Use the `branch_id` from the URL parameter
    batches = Batch.objects.filter()
    branch_stud = {}
    for batch in batches:
        count = Student.objects.filter(batch_id=batch.id).count()
        branch_stud[batch.batch] = count

    print(branch_stud)
    context = {'batches': batches, 'branch_stud': branch_stud,'branch_id':branch_id}
    print(context)
    return render(request, 'b.html', context)

@login_required
def sem(request,batch_id,branch_id):
    emp_id=request.session['empid'] 
    if not request.user.is_professor:
        messages.error(request, 'Please log in as a professor.')
        return redirect('professor_login')
    if not request.user.is_authenticated:
        return redirect('professor_login')
    if not emp_id:
        messages.error(request, 'Session expired. Please log in again.')
        return redirect('professor_login')
    # print(sem)
    sem=marks.objects.filter(batch_id=batch_id, branch_id=branch_id).values('sem').distinct()
   
    total_stud=Student.objects.filter(batch_id=batch_id, branch_id=branch_id).count()
    branch_stud = {}
    for s in sem:
        count = Document.objects.filter(batch_id=batch_id, branch_id=branch_id,sem=s['sem']).count()
        branch_stud[s['sem']] = count

    print(branch_stud)
    sem1=[]
    for s in sem:
        sem1.append(s['sem'])
    print(sem1)
    context = {'sem1': sem1, 'branch_stud': branch_stud,'total_stud':total_stud,'branch_id':branch_id,'batch_id':batch_id}
    return render(request, 'c.html', context)
# def view_students(request, sem):
#     """View students for a specific semester."""
#     students = Custo.objects.filter(sem=sem)  # Assuming you have a Student model
#     context = {'students': students, 'sem': sem}
#     return render(request, 'students.html', context)


def download(batch_id,branch_id,sem):
    wb=openpyxl.Workbook()
    ws=wb.active
    # resize image
    # img = Image.open('static/images/test.jpg')
    # img = img.resize((60, 60))
    # img.save('static/images/test.jpg')
    # img = openpyxl.drawing.image.Image('static/images/test.jpg')
    # img.anchor = 'A1'
    # ws.add_image(img)
    
    ws.cell(row=1, column=1).value = ''' JAIN COLLEGE OF ENGINEERING & RESEARCH, BELAGAVI
'''
    ws.cell(row=3, column=1).value = 'USN'
    ws.merge_cells('A3:A4')
    ws.cell(row=3, column=2).value = 'Student Name'
    ws.merge_cells('B3:B4')
    mark_list=['INT','EXT','TOTAL']
    subject_cell_list={}
    jump=0
    subject_list=[]
    # unique subject code
    user_marks=marks.objects.filter(batch_id=batch_id,branch_id=branch_id,sem=sem).values('subject_code').distinct()  # Get unique subject codes from marks.subject_code.unique()
    
    # user_marks=marks.objects.filter(batch_id=1,branch_id=1,sem=4).values('subject_code').distinct() 
    
    for i in user_marks:
       if i['subject_code'] not in subject_list:
            subject_list.append(i['subject_code'])
    
    for i in range(len(subject_list)):
        jump+=3
        ws.merge_cells(start_row=3, start_column=jump, end_row=3, end_column=jump+2)
        ws.cell(row=3, column=jump).value = subject_list[i]
        subject_cell_list[subject_list[i]]=(jump)
        for j in range(3):
            ws.cell(row=4, column=jump+j).value = mark_list[j]
            ws.cell(row=3, column=jump+j).alignment = Alignment(horizontal='center')
        # jump-=1
    # print(subject_cell_list)
    jump+=3
    ws.cell(row=3,column=jump).value='Result'
    ws.cell(row=3,column=jump+1).value='Total'
    ws.cell(row=3,column=jump+2).value='Percentage'
    ws.cell(row=3,column=jump+3).value='Class'

    

    # Enumerate the cells in the second row
    for cell in ws["3:3"]:
        cell.font = openpyxl.styles.Font.bold=True
    for cell in ws["4:4"]:
        cell.font = openpyxl.styles.Font.bold=True

    user_marks = marks.objects.select_related('usn').filter(batch_id=batch_id, branch_id=branch_id, sem=sem).values(
        'usn__username', 'name', 'subject_code', 'internal', 'external', 'total', 'result'
    ).order_by('usn__username')
    prev_usn=""
    row=4
    column=1
    flag=0
    sub_count=0
    analysis={}
    results={}
    class_analysis={'SC':0,'FC':0,'FCD':0,'Fail':0}
    for um in user_marks:
        if um['usn__username']!=prev_usn and flag==1:
            ws.cell(row=row,column=column+4).value=total
            ws.cell(row=row,column=column+3).value='Pass' if (total/sub_count)>40 else 'Fail'
            ws.cell(row=row,column=column+5).value=(total/sub_count)
            if (total/sub_count)>=40 and (total/sub_count)<60 :
                ws.cell(row=row,column=column+6).value='SC'
                class_analysis['SC']+=1
            elif (total/sub_count)>=60 and (total/sub_count)<80:
                ws.cell(row=row,column=column+6).value='FC'
                class_analysis['FC']+=1
            elif (total/sub_count)>=80 and (total/sub_count)<=100:
                ws.cell(row=row,column=column+6).value='FCD'
                class_analysis['FCD']+=1
            else:
                ws.cell(row=row,column=column+6).value='Fail'
                class_analysis['Fail']+=1
            if  (total/sub_count) in results:
                results[total/sub_count].append(prev_usn)
            else:
                results[total/sub_count]=[prev_usn]
            prev_usn=um['usn__username']
            total=0
            sub_count=0
            row+=1
            column=1
            ws.cell(row=row,column=column).value=um['usn__username']
            ws.cell(row=row,column=column+1).value=um['name']
        elif um['usn__username']!=prev_usn:
            row+=1
            ws.cell(row=row,column=column).value=um['usn__username']
            ws.cell(row=row,column=column+1).value=um['name']
            prev_usn=um['usn__username']
            total=0
            flag=1
            sub_count=0
        column=subject_cell_list[um['subject_code']]
        if um['subject_code'] not in analysis:
            analysis[um['subject_code']]=[1]
        else:
            analysis[um['subject_code']][0]+=1
        ws.cell(row=row,column=column).value=um['internal']
        ws.cell(row=row,column=column+1).value=um['external']
        ws.cell(row=row,column=column+2).value=um['total']
        
        if um['result']=='F'  and len(analysis[um['subject_code']])==1:
            analysis[um['subject_code']].append(1)
        elif um['result']=='F' and len(analysis[um['subject_code']])>1:
            analysis[um['subject_code']][1]+=1
        elif um['result']=='P' and len(analysis[um['subject_code']])==1:
            analysis[um['subject_code']].append(0)
        
        sub_count+=1
        try:
            total+=int(um['total'])
        except:
            total+=0
    ws.cell(row=row,column=column+4).value=total
    ws.cell(row=row,column=column+3).value='Pass' if (total/sub_count)>40 else 'Fail'
    ws.cell(row=row,column=column+5).value=(total/sub_count)
    if (total/sub_count)>=40 and (total/sub_count)<60 :
        ws.cell(row=row,column=column+6).value='SC'
        class_analysis['SC']+=1
    elif (total/sub_count)>=60 and (total/sub_count)<80:
        ws.cell(row=row,column=column+6).value='FC'
        class_analysis['FC']+=1
    elif (total/sub_count)>=80 and (total/sub_count)<=100:
        ws.cell(row=row,column=column+6).value='FCD'
        class_analysis['FCD']+=1
    else:
        ws.cell(row=row,column=column+6).value='Fail'
        class_analysis['Fail']+=1
    if  (total/sub_count) in results:
        results[total/sub_count].append(prev_usn)
    else:
        results[total/sub_count]=[prev_usn]
    # row+=2
    # ws.cell(row=row,column=2).value='Total Students'
    # ws.cell(row=row+1,column=2).value='Fail'
    # ws.cell(row=row+2,column=2).value='Pass'
    # ws.cell(row=row+3,column=2).value='Percentage'
    
    # for i in range(len(subject_list)):
    #     print(subject_cell_list[subject_list[i]])
    #     ws.cell(row=row,column=subject_cell_list[subject_list[i]]+2).value=analysis[subject_list[i]][0]
    #     ws.cell(row=row+1,column=subject_cell_list[subject_list[i]]+2).value=analysis[subject_list[i]][1]
    #     ws.cell(row=row+2,column=subject_cell_list[subject_list[i]]+2).value=analysis[subject_list[i]][0]-analysis[subject_list[i]][1]
    #     ws.cell(row=row+3,column=subject_cell_list[subject_list[i]]+2).value=((analysis[subject_list[i]][0]-analysis[subject_list[i]][1])/analysis[subject_list[i]][0])*100

    row+=5
    # ws.cell(row=row,column=2).value='FCD'
    # ws.cell(row=row+1,column=2).value='FC'
    # ws.cell(row=row+2,column=2).value='SC'
    # ws.cell(row=row+3,column=2).value='Fail'
    # ws.cell(row=row,column=3).value=class_analysis['FCD']
    # ws.cell(row=row+1,column=3).value=class_analysis['FC']
    # ws.cell(row=row+2,column=3).value=class_analysis['SC']
    # ws.cell(row=row+3,column=3).value=class_analysis['Fail']

    # row+=5
    # ws.cell(row=row,column=1).value='Toppers'
    # row+=1
    # ws.cell(row=row,column=2).value='USN'
    # ws.cell(row=row,column=3).value='Percentage'
    # rank=1
    # for i in reversed(sorted(results.keys())):
    #     ws.cell(row=row+rank,column=2).value=str(",".join(results[i]))
    #     ws.cell(row=row+rank,column=3).value=i
    #     rank+=1
    #     if rank>3:
    #         break
    import os
    batch=Batch.objects.get(id=batch_id)
    branch=Branch.objects.get(id=branch_id)
    ws.cell(row=2, column=1).value = "Semester :"+str(sem)
    ws.cell(row=2, column=4).value = "Branch :"+str(branch)
    ws.cell(row=2, column=7).value = "Batch :"+str(batch)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(subject_list)*3+6)
    ws['A1'].alignment = openpyxl.styles.alignment.Alignment(horizontal="center")
    ws['A1'].font = openpyxl.styles.Font.bold = True
    ws['A1'].font = openpyxl.styles.Font(size=20)

    

    # Set alignment for all cells
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            if cell.column_letter != 'B':  # Skip column B
                cell.alignment = Alignment(horizontal='center')
    for row in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            
            cell.border=openpyxl.styles.Border(
                left = Side(style='thin'),
                right = Side(style='thin'),
                top = Side(style='thin'),
                bottom = Side(style='thin')
            )

    

    os.makedirs('media/excel', exist_ok=True)
    save_path = 'media/excel/'+str(branch)+'_'+str(batch)+'_'+str(sem)+'.xlsx'
    wb.save(save_path)

    return save_path

@login_required
def download_report(request, batch_id, branch_id, sem):
    emp_id=request.session['empid'] 
    if not request.user.is_professor:
        messages.error(request, 'Please log in as a professor.')
        return redirect('professor_login')
    if not request.user.is_authenticated:
        return redirect('professor_login')
    if not emp_id:
        messages.error(request, 'Session expired. Please log in again.')
        return redirect('professor_login')
    # Generate the report
    file_path = download(batch_id, branch_id, sem)
    
    # Open the generated file and serve it to the user
    file = open(file_path, 'rb')  # Open the file in binary mode
    response = FileResponse(file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    batch=Batch.objects.get(id=batch_id)
    branch=Branch.objects.get(id=branch_id)
    response['Content-Disposition'] = f'attachment; filename="{branch}_{batch}_{sem}_report.xlsx"'
    
    return response


@login_required
def download_and_display(request, batch_id, branch_id, sem):
    emp_id=request.session['empid'] 
    if not request.user.is_professor:
        messages.error(request, 'Please log in as a professor.')
        return redirect('professor_login')
    if not request.user.is_authenticated:
        return redirect('professor_login')
    if not emp_id:
        messages.error(request, 'Session expired. Please log in again.')
        return redirect('professor_login')
    import openpyxl
    from django.shortcuts import render

    # Create workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Student Results"

    # Column headers for the main table
    ws.cell(row=3, column=1).value = 'USN'
    ws.merge_cells('A3:A4')
    ws.cell(row=3, column=2).value = 'Student Name'
    ws.merge_cells('B3:B4')

    mark_list = ['INT', 'EXT', 'TOTAL']
    subject_cell_list = {}
    subject_list = []


    # Get unique subject codes
    user_marks = marks.objects.filter(batch_id=batch_id, branch_id=branch_id, sem=sem).values('subject_code').distinct()
    for i in user_marks:
        if i['subject_code'] not in subject_list:
            subject_list.append(i['subject_code'])

    # Prepare column headers for subjects
    for i in range(len(subject_list)):
        ws.merge_cells(start_row=3, start_column=3 + i * 3, end_row=3, end_column=5 + i * 3)
        ws.cell(row=3, column=3 + i * 3).value = subject_list[i]
        subject_cell_list[subject_list[i]] = 3 + i * 3
        for j in range(3):
            ws.cell(row=4, column=3 + i * 3 + j).value = mark_list[j]

    # Fetch user marks and organize data
    user_marks = marks.objects.select_related('usn').filter(batch_id=batch_id, branch_id=branch_id, sem=sem).values(
        'usn__username', 'name', 'subject_code', 'internal', 'external', 'total', 'result'
    ).order_by('usn__username')

    print(user_marks)

    table_data = []
    others={}
    student_data = {}
    count=0
    subject_totals = {sub: {'total': 0, 'fail': 0, 'pass': 0} for sub in subject_list}
    class_stat={'FCD':0,'FC':0,'SC':0,'Fail':0}
    # Process marks for each student
    for um in user_marks:
        other={}
        if um['usn__username'] not in student_data:
            student_data[um['usn__username']] = {
                'name': um['name'],
                'subjects': {sub: {'internal': 0, 'external': 0, 'total': 0} for sub in subject_list},
                'total': 0,
                'percentage': 0,
                'result': '',
                'class': ''
            }
        student_data[um['usn__username']]['subjects'][um['subject_code']] = {
            'internal': um['internal'],
            'external': um['external'],
            'total': um['total']
        }
        try:
            if int(um['total']):
                student_data[um['usn__username']]['total'] += int(um['total'])
        except:
            student_data[um['usn__username']]['total'] += 0
        print(um['total'],um['subject_code'])
        subject_totals[um['subject_code']]['total'] += 1
        if um['result'] == 'F':
            subject_totals[um['subject_code']]['fail'] += 1
        elif um['result']=='P':
            subject_totals[um['subject_code']]['pass'] += 1
        else:
            other['usn']=um['usn__username']
            other['name']=um['name']
            other['subject_code']=um['subject_code']
            other['result']=um['result']
            others[count]=other
            count+=1

    # Calculate overall results
    for usn, data in student_data.items():
        data['percentage'] = data['total'] / len(subject_list)
        if data['percentage'] >= 80:
            data['class'] = 'FCD'
            class_stat['FCD']+=1
        elif data['percentage'] >= 60:
            data['class'] = 'FC'
            class_stat['FC']+=1
        elif data['percentage'] >= 40:
            data['class'] = 'SC'
            class_stat['SC']+=1
        else:
            data['class'] = 'Fail'
            class_stat['Fail']+=1
        data['result'] = 'Pass' if data['percentage'] >= 40 else 'Fail'
        table_data.append({
            'usn': usn,
            'name': data['name'],
            'subjects': data['subjects'],
            'total': data['total'],
            'percentage': data['percentage'],
            'result': data['result'],
            'class': data['class']
        })
        

    # Rank students by percentage
    ranked_students = sorted(table_data, key=lambda x: x['percentage'], reverse=True)
    ranks = [{'rank': i + 1, 'usn': student['usn'], 'name': student['name'], 'percentage': student['percentage']}
             for i, student in enumerate(ranked_students) if i < 3]


    for subject in subject_totals:
        subject_totals[subject]['percentage'] = (subject_totals[subject]['pass'] / subject_totals[subject]['total']) * 100

    overall_stats = {
        'subject_totals': subject_totals
    }
    subjects=[]
    pass_students=[]
    for subject in subject_totals:
        subjects.append(subject)
        pass_students.append(subject_totals[subject]['pass'])

    np_subjects=np.array(subjects)
    np_pass_students=np.array(pass_students)
    print(np_subjects)
    print(np_pass_students)
    plt.bar(np_subjects, np_pass_students)
    plt.xlabel('Subjects')
    plt.ylabel('Number of Pass Students')
    plt.title('Number of Pass Students in Each Subject')
    plt.xticks(rotation=45)
    plt.tight_layout()
    plt.savefig('static/images/bar_chart.png')
    plt.close()
    labels = []
    sizes = []

    for x, y in class_stat.items():
        labels.append(x)
        sizes.append(y)
    plt.pie(sizes,autopct='%1.1f%%')
    plt.legend(labels,)
    plt.savefig('static/images/pie_chart.png')
    plt.close()


        
    batch = Batch.objects.get(id=batch_id)
    branch = Branch.objects.get(id=branch_id)
    print(others)

    # Pass data to template
    return render(request, 'result.html', {
        'table_data': table_data,
        'subjects': subject_list,
        'ranks': ranks,
        'overall_stats': overall_stats,
        'class_stat':class_stat,
        'batch_id':batch_id,
        'branch_id':branch_id,
        'sem':sem,
        'branch':branch,
        'batch':batch,
        'others':others,
        
    })
    
@login_required
def logout_view(request):
    request.session.flush()
    messages.success(request, 'You have been logged out.')
    return redirect("index")










