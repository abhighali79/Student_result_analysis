from django.http import JsonResponse
from django.views.decorators.csrf import csrf_exempt
from .models import Document,marks
from django.shortcuts import render,redirect
import pandas as pd
import re ,os
from users.models import Branch,Batch,CustomUser, Student
import pandas as pd
import openpyxl
from django.contrib.auth import logout
from django.contrib import messages
from django.conf import settings
from django.contrib.auth.decorators import login_required




# def upload_page(request):
#     """Render the upload page."""
#     return render(request, 'upload.html')
df=pd.read_excel('rule.xlsx')

def parse_text(file):
    """
    Parses a text file to extract structured data based on labels and patterns.

    Args:
        file (str): Path to the file to be processed (without `.txt` extension).
        df (pd.DataFrame): DataFrame containing label definitions and rules.

    Returns:
        tuple: A dictionary of header data and a dictionary of line item data.
    """
    # Read the text file
    file_path = f"{str(file)[:-4]}.txt"
    with open(file_path, 'r', encoding='utf-8') as f:
        text = f.read()

    result_dict = {}
    line_item_start_indices = []
    line_item_dict = {}
    int_labels = {'Internal', 'External', 'Total'}
    text_lines = text.split('\n')

    # Iterate over each label definition in the DataFrame
    for idx, label_row in df.iterrows():
        label_key = label_row[0]
        label_type = label_row[1]
        label_names = label_row[2].split(',')

        found = False  # Track if a header match was found

        for line_no, line in enumerate(text_lines):
            for label in label_names:
                label = label.strip()

                # Process headers
                if label in line and label_type == 'header':
                    idx1 = line.find(label)
                    match = re.search(r":[\S\s]*", line[idx1 + len(label):])
                    if match:
                        result_dict[label_key] = match.group().replace(':', '').strip()
                        if label_key=='sem':
                            semester_label=label
                            semester=result_dict['sem']
                        found = True
                        break

                # Identify the start of line items
                    
                elif label in line and label_type == 'Line items' and not line_item_start_indices:
                    line_item_start = line_no
                    idx1 = line.find(label)
                    

                    # Collect all line items until a stopping condition
                    for line_no1 in range(line_no + 1, len(text_lines)):
                        line1 = text_lines[line_no1]
                        if "semester" in line1.lower():
                            sem_idx = line1.find(semester_label)
                            match = re.search(r"\d", line1[sem_idx + len(semester_label):])
                            print(match,line1[sem_idx + len(semester_label):])
                            if match:
                                semester=match.group().replace(':', '').strip()
                                print(semester)
                        if 'Abbreviations' in line1 :
                            break
                        if "results.vtu.ac.in" in line1:
                            continue

                        # Extract potential line items based on a pattern
                        string = line1[idx1-3:].strip().split(" ")[0].strip()
                        match = re.search(r"\S+\d{2,3}", string)
                        print(match,string)
                        # print(match,line_no1,string)
                        if match:
                            line_item_dict[line_no1] = [semester]
                            line_item_dict[line_no1].append(match.group(0).strip())
                            line_item_start_indices.append(line_no1)

            if found:
                break

        # Process additional line item data
        if line_item_start_indices and label_type == 'Line items' and label in text_lines[line_item_start]:

            if label_row[0]=='subject_code':
                continue
            
            idx1 = text_lines[line_item_start].find(label)

            for line_no2 in line_item_start_indices:
                string = text_lines[line_no2][idx1 - 1:idx1 + len(label) + 1]

                if label_row[0] == 'result':
                    if 'P' in string:
                        line_item_dict[line_no2].append('P')
                    elif 'F' in string:
                        line_item_dict[line_no2].append('F')
                    elif 'A' in string:
                        line_item_dict[line_no2].append('A')
                    elif 'W' in string:
                        line_item_dict[line_no2].append('W')
                    elif 'NE' in string:
                        line_item_dict[line_no2].append('NE')
                    elif 'X' in string:
                        line_item_dict[line_no2].append('X')
                    # match = re.search(r"\s*\s", string)
                    # if match:
                    #     line_item_dict[line_no2].append(match.group().strip())
                elif label in int_labels:
                    match = re.search(r"\S+\s*\S*", string)
                    if match:
                        line_item_dict[line_no2].append(match.group().strip())
                else:
                    match = re.search(r"\s*\d{2}\S*\s", string)
                    if match:
                        line_item_dict[line_no2].append(match.group().strip())

    return result_dict, line_item_dict
    



def insert_database(header, line_items, usn):
    # Fetch the Batch and Branch instances using the foreign keys from CustomUser
    student= Student.objects.get(usn=usn)
    batch = Batch.objects.get(id=student.batch_id)  # Fetch Batch instance
    branch = Branch.objects.get(id=student.branch_id)  # Fetch Branch instance
    user = CustomUser.objects.get(username=usn)  # Get the user with the given USN
    print(line_items)

    for item in line_items:
        # Create a `marks` record for each line item
        if marks.objects.filter(usn=user, subject_code=line_items[item][1],sem=line_items[item][0]).exists():
            #update
            marks.objects.filter(usn=user, subject_code=line_items[item][1],sem=line_items[item][0]).update(
                internal=line_items[item][2],
                external=line_items[item][3],
                total=line_items[item][4],
                result=line_items[item][5]
            )
        else:
            # print(item)
            marks.objects.create(
                usn=user,
                name=header['name'],
                sem=line_items[item][0],
                batch=batch,  # Pass the Batch instance
                branch=branch,  # Pass the Branch instance
                subject_code=line_items[item][1],
                internal=line_items[item][2],
                external=line_items[item][3],
                total=line_items[item][4],
                result=line_items[item][5]
            )

def pdf_to_text(file):
    # print("media\documents\\"+ str(file))
   # pdftotext.exe -layout -marginb 15 -enc UTF-8 media\documents\mech_ms.pdf      
    import subprocess
    command = [
            'pdftotext.exe',             # Executable           # Encoding (UTF-8)
            '-layout',  
            '-enc', 'UTF-8',                 # Maintain original layout
            '-table',                    # Generate a table layout
            '-fixed', '4',               # Use fixed width for columns
            '-nopgbrk',                  # Do not insert page breaks
             str(file)  ,      # Input PDF file path
             str(file)[:-4] + '.txt'  # Output text file path
        ]
    subprocess.run(command, check=True, stdout=subprocess.PIPE)
    

@login_required
@csrf_exempt
def upload_file(request):
    usn = request.session.get('usn')
    if not usn:
        messages.error(request, 'Session expired. Please log in again.')
        return redirect('student_login')

    user = CustomUser.objects.get(username=usn)
    if not user.is_student:
        messages.error(request, 'Please log in as a student.')
        return redirect('student_login')
    
    if usn:
        user_marks_sem=marks.objects.filter(usn=user.id).values('sem').distinct()
        user_marks = marks.objects.select_related('usn').filter(usn=user.id).values(
        'usn__username', 'name', 'subject_code', 'internal', 'external', 'total', 'result','sem'
    ).order_by('usn__username')
        # user_marks=marks.objects.filter(usn=usn)
        # print(user_marks)
        sem_percentage={}
        for semester in user_marks_sem:
           
            total=0
            subject_count=0
            for um in user_marks:
                
                if um['sem']==semester['sem']:
                    subject_count+=1
                    try:
                        if int(um['total']):
                            total+=int(um['total'])
                    except:
                        total+=0
            sem_percentage[semester['sem']]=total/subject_count
        # print(sem_percentage)
            
    



    """Handle file upload."""
    if request.method == 'POST' and request.FILES.get('file'):
        usn = request.session.get('usn')
        if not usn:
            return render(request, 'upload.html', {'error': 'Session expired. Please log in again.'})

        file = request.FILES['file']
        title = request.POST.get('title', '')
        
        # Save the file temporarily
        temp_dir = os.path.join(settings.BASE_DIR, 'temp')
        os.makedirs(temp_dir, exist_ok=True)  # Create temp directory if it doesn't exist
        temp_path = os.path.join(temp_dir, file.name.replace(" ", "_"))

        with open(temp_path, 'wb+') as destination:
            for chunk in file.chunks():
                destination.write(chunk)

        # Process the file (e.g., PDF to text, parsing)
        pdf_to_text(temp_path)
        header, line_items = parse_text(temp_path)
        print(header,line_items)
        if header['usn'] != usn:
            messages.error(request, 'USN does not match the uploaded file.')
            return render(request, 'index.html')
        # print( CustomUser.objects.get(usn=header['usn']))
        # Insert into the database

        # Move file to final destination (MEDIA_ROOT)
        final_dir = os.path.join(settings.MEDIA_ROOT, 'uploads')
        os.makedirs(final_dir, exist_ok=True)
        final_path = os.path.join(final_dir, usn + "_" + header['sem']+".pdf")
        if os.path.exists(final_path):
            messages.error(request, 'File already exists.')
            return render(request, 'index.html')
        os.rename(temp_path, final_path)
        # check file is present or not

        insert_database(header, line_items, CustomUser.objects.get(username=header['usn']).username)
        # Save file metadata to the database
        Document.objects.create(
            usn=CustomUser.objects.get(username=usn),
            sem=header['sem'],  # Ensure 'sem' is correctly extracted from header
            branch=Branch.objects.get(id=Student.objects.get(usn=usn).branch_id),
            batch=Batch.objects.get(id=Student.objects.get(usn=usn).batch_id),
            file=os.path.join('uploads', final_path)
        )
        messages.success(request, 'File uploaded successfully!')
        return render(request, 'index.html',{'sem_percentage': sem_percentage})
    # messages.error(request, 'No file uploaded.')
    return render(request, 'index.html',{'sem_percentage': sem_percentage})


@login_required
def display_table(request):
    usn = request.session.get('usn')
    if not usn:
        messages.error(request, 'Session expired. Please log in again.')
        return redirect('student_login')

    user = CustomUser.objects.get(username=usn)
    if not user.is_student:
        messages.error(request, 'Please log in as a student.')
        return redirect('student_login')
    
    usn = request.session.get('usn') 
    documents=Document.objects.filter(usn=usn)
    return render(request, 'uploaded_documenets.html', {'documents': documents})


# @login_required
def logout_view(request):
   
    request.session.flush()
    messages.success(request, 'You have been logged out.')
    return redirect('index')

# @csrf_exempt
# def download(request):
#     wb=openpyxl.Workbook()
#     ws=wb.active
#     ws.cell(row=3, column=1).value = 'USN'
#     ws.merge_cells('A3:A4')
#     ws.cell(row=3, column=2).value = 'student_name'
#     ws.merge_cells('B3:B4')
#     mark_list=['INT','EXT','TOTAL']
#     subject_cell_list={}
#     jump=0
#     subject_list=[]
#     # unique subject code
#     # user_marks=marks.objects.filter(batch_id=request.session.get('batch_id'),branch_id=request.session.get('branch_id'),sem=request.session.get('sem')).order_by('subject_code').values('subject_code').distinct()  # Get unique subject codes from marks.subject_code.unique()
    
#     user_marks=marks.objects.filter(batch_id=1,branch_id=1,sem=4).values('subject_code').distinct() 
    
#     for i in user_marks:
#        if i['subject_code'] not in subject_list:
#             subject_list.append(i['subject_code'])
    
#     for i in range(len(subject_list)):
#         jump+=3
#         ws.merge_cells(start_row=3, start_column=jump, end_row=3, end_column=jump+2)
#         ws.cell(row=3, column=jump).value = subject_list[i]
#         subject_cell_list[subject_list[i]]=(jump)
#         for j in range(3):
#             ws.cell(row=4, column=jump+j).value = mark_list[j]
#         # jump-=1
#     # print(subject_cell_list)
#     jump+=3
#     ws.cell(row=3,column=jump).value='Result'
#     ws.cell(row=3,column=jump+1).value='Total'
#     ws.cell(row=3,column=jump+2).value='Percentage'
#     ws.cell(row=3,column=jump+3).value='Class'

#     user_marks=marks.objects.filter(batch_id=1,branch_id=1,sem=4).order_by('usn').values('usn','name','subject_code','internal','external','total','result')
#     print(user_marks)
#     prev_usn=""
#     row=4
#     column=0
#     flag=0
#     sub_count=0
#     for um in user_marks:
#         if um['usn']!=prev_usn and flag==1:
#             ws.cell(row=row,column=column+4).value=total
#             ws.cell(row=row,column=column+3).value='Pass' if (total/sub_count)*100>40 else 'Fail'
#             ws.cell(row=row,column=column+5).value=(total/sub_count)*100
#             ws.cell(row=row,column=column+3).value='Pass' if (total/sub_count)*100>40 else 'Fail'
#             prev_usn=um['usn']
#             total=0
#             sub_count=0
#             row+=1
#         elif um['usn']!=prev_usn:
#             prev_usn=um['usn']
#             total=0
#             sub_count=0
#             row+=1
#         column=subject_cell_list[um['subject_code']]
#         ws.cell(row=row,column=column).value=um['internal']
#         ws.cell(row=row,column=column+1).value=um['external']
#         ws.cell(row=row,column=column+2).value=um['total']
#         sub_count+=1
#         total+=um['total']
#     ws.cell(row=row,column=column+4).value=total
#     ws.cell(row=row,column=column+3).value='Pass' if (total/sub_count)>40 else 'Fail'
#     ws.cell(row=row,column=column+5).value=(total/sub_count)
#     ws.cell(row=row,column=column+3).value='Pass' if (total/sub_count)>40 else 'Fail'
            

#     save_path = 'result1.xlsx'
#     wb.save(save_path)
#     # for i in a['subject_code'].values():
#     #     if i not in subject_list:
#     #         subject_list.append(i)

#     # for i in user:
#     #     print(i.usn)
#     #     print(i.name)
#     return render(request, 'index.html')
